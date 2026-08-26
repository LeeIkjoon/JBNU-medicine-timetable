# -*- coding: utf-8 -*-
"""계명대 의대 2026-2학기 강의시간표 xlsx → 앱 페이로드 변환.

포맷: 학년별 시트(의예1/의예2/의학1/의학2/의학3), 행 = 주차/일시/교시/시작시간/
교과목명/교수명/수업주제/강의실. 'N주차' 구분행·빈 과목행은 제외.
시험 판정은 수업주제 기준(중간·기말고사, N차 시험, 정기시험, 모의고사) —
'평가'는 일반 주제에 흔해 제외.

사용: python3 tools/build_kmu_2026_2.py <xlsx경로> <출력디렉토리>
"""
import sys, os, re, json, time
from datetime import date
import openpyxl

PERIOD_START = {1:'8:30',2:'9:30',3:'10:30',4:'11:30',5:'13:30',6:'14:30',7:'15:30',8:'16:30',9:'17:30',10:'18:30'}
PERIOD_END   = {1:'9:20',2:'10:20',3:'11:20',4:'12:20',5:'14:20',6:'15:20',7:'16:20',8:'17:20',9:'18:20',10:'19:20'}
GRADE_LABELS = {'의예1':'의예과 1학년','의예2':'의예과 2학년','의학1':'의학과 1학년',
                '의학2':'의학과 2학년','의학3':'의학과 3학년','의학4':'의학과 4학년'}
WKN = ['월','화','수','목','금','토','일']
EXAM_RE = re.compile(r'중간\s*(고사|시험)|기말\s*(고사|시험)|\d\s*차\s*시험|정기\s*시험|모의고사|^\s*시험\s*$')

def main(xlsx_path, out_dir):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    from collections import Counter
    for sheet in wb.sheetnames:
        label = GRADE_LABELS.get(sheet.strip())
        if not label:
            continue
        ws = wb[sheet]
        items, wdd = [], {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8):
            v = [c.value for c in row]
            subj = str(v[4]).strip() if v[4] else ''
            if not subj or re.fullmatch(r'\d+주차', subj):
                continue
            if v[0] is None or v[1] is None or v[2] is None:
                continue
            try:
                week = str(int(v[0]))
                period = int(v[2])
            except (TypeError, ValueError):
                continue
            ds = str(v[1])[:10]
            y, m, dd = map(int, ds.split('-'))
            day = WKN[date(y, m, dd).weekday()]
            prof = str(v[5]).strip() if v[5] else ''
            topic = ' '.join(str(v[6]).split()) if v[6] else ''
            room = str(v[7]).strip() if v[7] else ''
            is_exam = bool(EXAM_RE.search(topic)) or bool(EXAM_RE.search(subj))
            it = {'week': week, 'date': ds, 'day': day, 'period': period,
                  'start': PERIOD_START.get(period, ''), 'end': PERIOD_END.get(period, ''),
                  'subject': subj, 'professor': prof, 'is_exam': is_exam}
            if topic:
                it['topic'] = topic
            if room and room != '추후안내':
                it['room'] = room
            items.append(it)
            wdd.setdefault(week, {})[day] = ds
        if not items:
            continue
        ed = sorted({i['date'] for i in items if i['is_exam']})
        wks = sorted({i['week'] for i in items}, key=int)
        ts = int(time.time() * 1000)
        payload = {'items': items, 'wdd': wdd, 'ed': ed, 'wks': wks,
                   'grade': label, 'ts': ts,
                   'changelog': {'ts': ts, 'msg': '2026-2학기 시간표 적용'}}
        key = 'kmu_' + label.replace(' ', '_')
        out = os.path.join(out_dir, key + '.json')
        with open(out, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=1)
        cnt = Counter(i['subject'] for i in items)
        print(f'== {label} → {out}')
        print('  items:', len(items), '| weeks:', wks[0], '~', wks[-1], '| 시험일:', len(ed))
        for s, n in cnt.most_common(6):
            nprof = len({i['professor'] for i in items if i['subject'] == s and i['professor']})
            print(f'  {n:4d}  {s[:26]}  (교수 {nprof}명)')

if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])

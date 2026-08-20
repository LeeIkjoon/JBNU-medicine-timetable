# -*- coding: utf-8 -*-
"""2026-2학기 의예과 2학년 시간표 xlsx → 앱 Firebase 페이로드(JSON) 변환.

- 그리드: '예2 전체시간표(해부학분반적용)' 시트 (화/수/금 실습 분반 전체 포함;
  앱의 분반 선택 기능이 렌더 시 본인 분반만 표시)
- 교수: '인체육안구조'(이론) / '인체육안구조실습' 수업계획표 시트를
  (날짜, 교시)로 매칭해 병합. 실습 계획표는 수요일 기준이라 화/금 실습은 교수 공란.
- 실습 계획표의 강의주제에 test가 들어간 교시는 is_exam 처리(땡시).

사용: python3 tools/build_premed2_2026_2.py <xlsx경로> <출력.json>
"""
import sys, json, time
import openpyxl

PERIOD_START = {1:'8:30',2:'9:30',3:'10:30',4:'11:30',5:'13:30',6:'14:30',7:'15:30',8:'16:30',9:'17:30',10:'18:30'}
PERIOD_END   = {1:'9:20',2:'10:20',3:'11:20',4:'12:20',5:'14:20',6:'15:20',7:'16:20',8:'17:20',9:'18:20',10:'19:20'}

def ymd(v):
    return str(v)[:10]

def load_plan(ws):
    """수업계획표 시트 → {(date, period): (professor, topic)}"""
    out = {}
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, max_col=8):
        v = [c.value for c in row]
        if v[1] is None or v[3] is None:
            continue
        try:
            period = int(v[3])
        except (TypeError, ValueError):
            continue
        prof = str(v[6]).strip() if v[6] else ''
        topic = str(v[7]).strip() if v[7] else ''
        out[(ymd(v[1]), period)] = (prof, topic)
    return out

def main(xlsx_path, out_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    theory_plan = load_plan(wb['인체육안구조'])
    prac_plan   = load_plan(wb['인체육안구조실습'])

    ws = wb['예2 전체시간표(해부학분반적용)']
    items, wdd = [], {}
    week = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=13):
        v = [c.value for c in row]
        if v[0] is not None:
            week = str(v[0]).strip()
        if v[1] is None or week is None:
            continue
        date, day = ymd(v[1]), str(v[2]).strip()
        wdd.setdefault(week, {})[day] = date
        for col in range(3, 13):
            if v[col] is None:
                continue
            subject = ' '.join(str(v[col]).split())
            if not subject:
                continue
            period = col - 2
            prof, is_exam = '', False
            if subject.startswith('인체육안구조('):
                prof = theory_plan.get((date, period), ('', ''))[0]
            elif subject == '인체육안구조실습':
                prof, topic = prac_plan.get((date, period), ('', ''))
                if 'test' in topic.lower():
                    is_exam = True
            items.append({
                'week': week, 'date': date, 'day': day, 'period': period,
                'start': PERIOD_START[period], 'end': PERIOD_END[period],
                'subject': subject, 'professor': prof, 'is_exam': is_exam,
            })

    ed = sorted({it['date'] for it in items if it['is_exam']})
    wks = sorted({it['week'] for it in items}, key=int)
    ts = int(time.time() * 1000)
    payload = {
        'items': items, 'wdd': wdd, 'ed': ed, 'wks': wks,
        'grade': '의예과 2학년', 'ts': ts,
        'changelog': {'ts': ts, 'msg': '2026-2학기 시간표 적용'},
    }
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)

    # 검증 요약
    from collections import Counter
    cnt = Counter(it['subject'] for it in items)
    print('items:', len(items), '| weeks:', wks[0], '~', wks[-1], '| exam dates:', ed)
    for s, n in cnt.most_common():
        profs = Counter(it['professor'] or '(없음)' for it in items if it['subject'] == s)
        print(f'  {n:4d}  {s}  {dict(profs)}')

if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
